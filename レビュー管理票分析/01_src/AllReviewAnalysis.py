# -*- coding: utf-8 -*-

import openpyxl 
import glob
import time
from copy import copy

# 定義
max_r = 14 # 最大の行
max_l = 66 # 最大の列

calc_r = [2,3,4,5,6,7,8,9,10,11,12,13] # 合計計算対象行
calc_l_tmp = [3,4,5] # 合計計算対象列
calc_l = [3,4,5]

for i in range(10):# 合計計算対象拡張
	calc_l_tmp = list(map(lambda x: x+6, calc_l_tmp))
	calc_l.extend(calc_l_tmp)

measure_sheet_name = 'ドキュメント指摘計測'
outputfile_name = 'AllReviewAnalysis_Rst.xlsx'

# コンフィグ読む
path = r'config/cfg_AllReviewAnalysis.txt'
with open(path, encoding='shift-jis') as f:
	ReadFileList = [s.strip() for s in f.readlines()]
# コンフィグ結果出力
print('ReadFileList')
print(ReadFileList)
# 合計配列作成
arr = [[0 for i in range(max_l)] for j in range(max_r)]

# 全エクセルに対して処理
for wb_name in ReadFileList:
	print('processing...'+wb_name)
	wb=openpyxl.load_workbook(wb_name, data_only=True) # 数式でなく計算結果を取得する
	# python4
	#sheet = wb.get_sheet_by_name(sheet_name)
	# python3
	sheet = wb[measure_sheet_name]
	# 各セル合計値算出
	for i in calc_r:
		for j in calc_l:
			if sheet.cell(row=i, column=j).value:
					arr[i][j] = arr[i][j] + sheet.cell(row=i, column=j).value

# エクセル作成
new_wb = wb=openpyxl.load_workbook(outputfile_name, data_only=False) # 数式でなく計算結果を取得する

# ドキュメント指摘計測 シート作成
new_measure_sheet = wb[measure_sheet_name]

# 合計値代入
for i in calc_r:
	for j in calc_l:
		if (i == 1) or (j % 6 != 5):
			new_measure_sheet.cell(row=i, column=j).value = arr[i][j]

# 保存
new_wb.save(outputfile_name)

# Finish
print('Finish!')

# wait
print('wait 30s...')
time.sleep(30)
#print(glob.glob("*.xlsx"))