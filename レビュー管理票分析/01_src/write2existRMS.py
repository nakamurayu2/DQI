# -*- coding: utf-8 -*-

import os
import openpyxl 
import glob
import time
from copy import copy
import xlwings as xw
import re
import win32com.client as win32
import shutil

# 定義
work_dir_name = 'work'

# コンフィグ読む
print('Reading Config...')
path = r'config\cfg_write2existRMS.txt'

with open(path,'r', encoding='shift-jis') as f: # 
	ReadFileList = [s.strip() for s in f.readlines()]

cfg_mode = 0 # 0:通常,1：PERSON,2:TARGET_FOLDER 
PERSON = list() # 対象の人
TARGET_FOLDER = list() # 対象フォルダ
CONDITION = '' # 指摘・不具合の条件式
for line in ReadFileList:
	if re.search(r'###\s?\[PERSON\]', line):
		cfg_mode = 1
	elif re.search(r'###\s?\[TARGET_FOLDER\]', line):
		cfg_mode = 2
	elif re.search(r'###\s?\[CONDITION\]', line):
		cfg_mode = 3		
	elif re.search(r'^#', line):
		continue
	elif cfg_mode == 1:
		PERSON.append(line)
	elif cfg_mode == 2:
		TARGET_FOLDER.append(line)
	elif cfg_mode == 3:
		CONDITION = line			
	else:
		cfg_mode = 0

print('PERSON')
print(PERSON)
print('TARGET_FOLDER')
print(TARGET_FOLDER)
print('CONDITION')
print(CONDITION)

# 対象フォルダからレビュー管理票検索
print('Search rvm File...')
rvm_list = list()
for folder in TARGET_FOLDER:
	print('Search ' + folder + '...')
	print('※※※時間かかります※※※')
	for p in glob.glob(folder + '/**/*レビュー管理表*.xlsm', recursive=True):
		if os.path.isfile(p):
			if 'テンプレート' in p:
				continue
			elif re.search(r'^~', os.path.basename(p)): # 編集中のやつ
				continue
			else:
				rvm_list.append(p)
				print(p)

# テンプレート開く
tmpl_wb=openpyxl.load_workbook(r'template\ピアレビュー管理表テンプレート.xlsm', data_only=False, keep_vba=True)
tmpl_sht1 = tmpl_wb['ドキュメント指摘計測']

print('PERSON')
print(PERSON)
print('rvm_list')
print(rvm_list)

# 作業フォルダ
shutil.rmtree(work_dir_name, ignore_errors=True)
os.makedirs(work_dir_name, exist_ok=True) # workフォルダ作成

wb_num = 0
new_wb_path_list = list()
# 全エクセルに対して処理
for wb_path in rvm_list:
	wb_num = wb_num + 1 
	print('processing...'+wb_path)
	# エクセルオープン
	wb=openpyxl.load_workbook(wb_path, data_only=True) #  keep_vba=True
	new_wb = openpyxl.Workbook()
	### レビュー管理シート
	# python4
	#sheet = wb.get_sheet_by_name(sheet_name)
	# python3
	print('レビュー管理シート...')
	rvm_sht = wb['レビュー管理']
	new_wb.create_sheet(title='レビュー管理')
	new_rvm_sht = new_wb['レビュー管理']
	# 各セルコピー
	for i in range(10, rvm_sht.max_row, 1):
		for j in range(7, 16, 1):
	 		new_rvm_sht.cell(row=i, column=j).value = rvm_sht.cell(row=i, column=j).value

	# 有効なシートのリストを作成
	valid_sht_idx_list = list()
	# (11,12...,35)に「=IF(ISERROR(B11)=TRUE,"",INDIRECT(CONCATENATE("'記録","#",$C11,"'!V$4")))」を書き込む
	for i in range(11, rvm_sht.max_row, 1):
		if rvm_sht.cell(row=i, column=2).value == '承認済': # 承認済みのものだけ対象
			if 'ALL' in PERSON or rvm_sht.cell(row=i, column=11).value in PERSON:
				# 有効なシートとして追加
				valid_sht_idx_list.append(i-10)
				# 数式書き込み 
				new_rvm_sht.cell(row=i, column=35).value = '=\'記録#' + str(i-10) + '\'!V4'
			else: # 条件にマッチしない
				new_rvm_sht.cell(row=i, column=35).value = '=0' # 0で上書き
	
	### 記録#tシート
	print('記録#tシート...')
	print(valid_sht_idx_list)
	for i in valid_sht_idx_list:
		rc_sht_name = '記録#'+str(i)
		rc_sht = wb[rc_sht_name]
		new_wb.create_sheet(title=rc_sht_name)
		new_rc_sht = new_wb[rc_sht_name]
		# 各セルコピー
		for i in range(48, rc_sht.max_row, 1): # openpyxlの不具合？一部数式でエラーが出るので
			for j in range(8, 10, 1):
				new_rc_sht.cell(row=i, column=j).value = rc_sht.cell(row=i, column=j).value
		# (4,22)に「=COUNTIF(H50:I200,"[文章]*")」を書き込む
		new_rc_sht.cell(row=4, column=22).value = CONDITION

	### ドキュメント指摘計測シート
	# シート作成
	new_ws = new_wb.create_sheet(title='ドキュメント指摘計測')
	# 各セルコピー
	for i in range(1, tmpl_sht1.max_row, 1):
		for j in range(1, tmpl_sht1.max_column, 1):
			new_ws.cell(row=i, column=j).value = tmpl_sht1.cell(row=i, column=j).value

	### 保存＆クローズ
	name, ext = os.path.splitext(os.path.basename(wb_path))
	new_wb_name = 'D_'+str(wb_num)+'_'+ name + '.xlsx'
	new_wb_path = work_dir_name + '\\' + new_wb_name
	new_wb.save(new_wb_path) # 保存

	### new_wb_path_listに追加
	new_wb_path_list.append(new_wb_path)

	### エクセルの数式反映処置(エクセルを完全にオープン・保存・クローズする)
	# Excelでxlsxファイルを保存すると、数式のセルは数式と値の両方がファイルに書き込まれますが、openpyxlでは数式しか書き込まれません。もしopenpyxlで数式を書き込んだファイルからそのまま値を読み出すとNoneが読み出されます。そのため、openpyxlで数式を書いたファイルから値を読み出すには、一度Excelでファイルを開いて保存する必要があります。
	app = xw.App()
	wb = app.books.open(new_wb_path)
	wb.save(path=None)
	wb.close()
	app.quit()
	# excel = win32.gencache.EnsureDispatch("Excel.Application")
	# p = pathlib.Path(new_wb_path)
	# workbook = excel.Workbooks.Open(p.resolve())
	# # this must be the absolute path (r"C:/abc/def/ghi")
	# workbook.Save()
	# workbook.Close()
	# excel.Quit()

# ログ作成
path = 'log_write2existRMS.txt'
f = open(path, mode='w')
f.write('\n'.join(new_wb_path_list))
f.close()

# Finish
print('Collect All Files Finish!')

# wait
# print('wait 30s...')
# time.sleep(30)
#print(glob.glob("*.xlsx"))

def listup_files(path):
    yield [os.path.abspath(p) for p in glob.glob(path, recursive=True)]